"""Build DSA reference spreadsheet with LeetCode problems column."""
import sys, os, json
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

DATA = json.load(open(os.path.join(os.path.dirname(__file__), 'dsa_enriched.json')))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = sys.argv[1] if len(sys.argv) > 1 else "DSA_Reference_Catalog.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Catalog"

HEADERS = [
    "Name", "Kind", "Category",
    "Definition",
    "Time complexity", "Space complexity",
    "Key techniques",
    "Wikipedia",
    "LeetCode problems",
    "Learning resources",
    "Sean Prashad", "Grind 75",
    "Notes",
]
ws.append(HEADERS)

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="2F5597")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)

for col_idx in range(1, len(HEADERS) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border

yes_fill = PatternFill("solid", start_color="C6EFCE")
no_fill = PatternFill("solid", start_color="FFEB9C")
kind_colors = {
    "Data Structure": "D9E1F2",
    "Algorithm": "FCE4D6",
    "Technique": "E2EFDA",
    "Paradigm": "FFF2CC",
}

for entry in DATA:
    lc_text = " | ".join(
        f"#{p['id']} {p['title']} ({p['difficulty'][0]})"
        for p in entry["leetcode"]
    )
    row = [
        entry["name"],
        entry["kind"],
        entry["category"],
        entry["definition"],
        entry["time"],
        entry["space"],
        entry["techniques"],
        entry["wikipedia"],
        lc_text,
        entry["resources"],
        "Yes" if entry["sean_prashad"] else "No",
        "Yes" if entry["grind75"] else "No",
        entry["notes"],
    ]
    ws.append(row)
    r = ws.max_row

    kc = ws.cell(row=r, column=2)
    if entry["kind"] in kind_colors:
        kc.fill = PatternFill("solid", start_color=kind_colors[entry["kind"]])
    sp = ws.cell(row=r, column=11)
    sp.fill = yes_fill if entry["sean_prashad"] else no_fill
    gr = ws.cell(row=r, column=12)
    gr.fill = yes_fill if entry["grind75"] else no_fill
    wiki_cell = ws.cell(row=r, column=8)
    if entry["wikipedia"]:
        wiki_cell.hyperlink = entry["wikipedia"]
        wiki_cell.font = Font(name="Calibri", color="0563C1", underline="single", size=10)
    # Hyperlink first LeetCode problem (col 9). Excel only allows one hyperlink per cell,
    # so we link the FIRST problem and the rest are visible as text.
    if entry["leetcode"]:
        lc_cell = ws.cell(row=r, column=9)
        lc_cell.hyperlink = entry["leetcode"][0]["url"]
        lc_cell.font = Font(name="Calibri", color="0563C1", underline="single", size=10)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=r, column=col)
        cell.alignment = left
        cell.border = border
        if cell.font.name != "Calibri":
            cell.font = Font(name="Calibri", size=10)

# Column widths
widths = [28, 16, 18, 60, 26, 18, 48, 36, 80, 48, 14, 12, 50]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

last_col = get_column_letter(len(HEADERS))
last_row = ws.max_row
tab = Table(displayName="DSACatalog", ref=f"A1:{last_col}{last_row}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws.add_table(tab)

# ---------------- LeetCode Problems sheet (one row per problem) ----------------
lc_sheet = wb.create_sheet("LeetCode Problems")
LC_HEADERS = ["Entry", "Kind", "Category", "#", "Problem", "Difficulty", "Acceptance %", "Premium", "G75", "B75", "NC150", "Done?", "URL"]
lc_sheet.append(LC_HEADERS)
for col_idx in range(1, len(LC_HEADERS) + 1):
    c = lc_sheet.cell(row=1, column=col_idx)
    c.font = header_font; c.fill = header_fill; c.alignment = center

diff_fills = {
    "Easy": PatternFill("solid", start_color="D4EDDA"),
    "Medium": PatternFill("solid", start_color="FFF3CD"),
    "Hard": PatternFill("solid", start_color="F8D7DA"),
}
for entry in DATA:
    for p in entry["leetcode"]:
        row = [
            entry["name"],
            entry["kind"],
            entry["category"],
            int(p["id"]) if p["id"].isdigit() else p["id"],
            p["title"],
            p["difficulty"],
            p["ac_rate"],
            "Yes" if p["premium"] else "No",
            "Yes" if p.get("grind75") else "",
            "Yes" if p.get("blind75") else "",
            "Yes" if p.get("neetcode150") else "",
            "",  # User fills in Done?
            p["url"],
        ]
        lc_sheet.append(row)
        r = lc_sheet.max_row
        # Difficulty colour
        df_cell = lc_sheet.cell(row=r, column=6)
        df_cell.fill = diff_fills.get(p["difficulty"], PatternFill())
        # Premium colour
        if p["premium"]:
            lc_sheet.cell(row=r, column=8).fill = PatternFill("solid", start_color="FFE5B4")
        # Badge colours
        if p.get("grind75"):
            lc_sheet.cell(row=r, column=9).fill = PatternFill("solid", start_color="C6F6D5")
        if p.get("blind75"):
            lc_sheet.cell(row=r, column=10).fill = PatternFill("solid", start_color="BEE3F8")
        if p.get("neetcode150"):
            lc_sheet.cell(row=r, column=11).fill = PatternFill("solid", start_color="E9D8FD")
        # URL hyperlink
        url_cell = lc_sheet.cell(row=r, column=13)
        url_cell.hyperlink = p["url"]
        url_cell.font = Font(name="Calibri", color="0563C1", underline="single", size=10)
        # All cells
        for col in range(1, len(LC_HEADERS) + 1):
            cell = lc_sheet.cell(row=r, column=col)
            cell.alignment = left
            if cell.font.name != "Calibri":
                cell.font = Font(name="Calibri", size=10)

lc_widths = [28, 16, 18, 8, 48, 11, 13, 10, 7, 7, 8, 8, 50]
for i, w in enumerate(lc_widths, start=1):
    lc_sheet.column_dimensions[get_column_letter(i)].width = w
lc_sheet.row_dimensions[1].height = 30
lc_sheet.freeze_panes = "A2"

last_col_lc = get_column_letter(len(LC_HEADERS))
last_row_lc = lc_sheet.max_row
lc_tab = Table(displayName="LeetCodeProblems", ref=f"A1:{last_col_lc}{last_row_lc}")
lc_tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
lc_sheet.add_table(lc_tab)

# ---------------- Progress Dashboard sheet ----------------
dash = wb.create_sheet("Progress Dashboard")
dash.append(["DSA Practice Progress Dashboard"])
dash["A1"].font = Font(name="Calibri", bold=True, size=18, color="2F5597")
dash.merge_cells("A1:E1")
dash.append([])
dash.append([
    "Mark problems as done in the 'LeetCode Problems' sheet by typing Y or any character in the 'Done?' column. "
    "This dashboard counts them automatically."
])
dash.merge_cells("A3:E3")
dash["A3"].alignment = Alignment(wrap_text=True, vertical="top")
dash["A3"].font = Font(name="Calibri", italic=True, color="6A737D")
dash.row_dimensions[3].height = 30
dash.append([])

# Totals
dash.append(["Totals", "", "", "", ""])
dash["A5"].font = Font(name="Calibri", bold=True, size=14)
dash.append(["Total problems", f"=COUNTA('LeetCode Problems'!E2:E{last_row_lc})"])
dash.append(["Completed",      f"=COUNTA('LeetCode Problems'!L2:L{last_row_lc})"])
dash.append(["% complete",     f"=IFERROR(B7/B6, 0)"])
dash["B8"].number_format = "0.0%"
dash["A6"].font = Font(bold=True)
dash["A7"].font = Font(bold=True)
dash["A8"].font = Font(bold=True, color="2F5597")

dash.append([])

# By difficulty
dash.append(["By difficulty", "Total", "Completed", "%"])
for c in range(1, 5):
    dash.cell(row=dash.max_row, column=c).font = Font(name="Calibri", bold=True)
    dash.cell(row=dash.max_row, column=c).fill = PatternFill("solid", start_color="2F5597")
    dash.cell(row=dash.max_row, column=c).font = Font(name="Calibri", bold=True, color="FFFFFF")

difficulties = ["Easy", "Medium", "Hard"]
for diff in difficulties:
    dash.append([
        diff,
        f'=COUNTIF(\'LeetCode Problems\'!F2:F{last_row_lc}, "{diff}")',
        f'=COUNTIFS(\'LeetCode Problems\'!F2:F{last_row_lc}, "{diff}", \'LeetCode Problems\'!L2:L{last_row_lc}, "<>")',
        f"=IFERROR(C{dash.max_row + 1}/B{dash.max_row + 1}, 0)",
    ])
    dash.cell(row=dash.max_row, column=4).number_format = "0.0%"

dash.append([])

# By curated list (G75 / B75 / NC150)
dash.append(["By curated list", "Total", "Completed", "%"])
header_row_cur = dash.max_row
for c in range(1, 5):
    dash.cell(row=header_row_cur, column=c).font = Font(name="Calibri", bold=True, color="FFFFFF")
    dash.cell(row=header_row_cur, column=c).fill = PatternFill("solid", start_color="2F5597")

curated_lists = [
    ("Grind 75",     "I"),
    ("Blind 75",     "J"),
    ("NeetCode 150", "K"),
]
for label, col_letter in curated_lists:
    dash.append([
        label,
        f'=COUNTIF(\'LeetCode Problems\'!{col_letter}2:{col_letter}{last_row_lc}, "Yes")',
        f'=COUNTIFS(\'LeetCode Problems\'!{col_letter}2:{col_letter}{last_row_lc}, "Yes", \'LeetCode Problems\'!L2:L{last_row_lc}, "<>")',
        f"=IFERROR(C{dash.max_row + 1}/B{dash.max_row + 1}, 0)",
    ])
    dash.cell(row=dash.max_row, column=4).number_format = "0.0%"

dash.append([])

# By category
dash.append(["By category", "Total", "Completed", "%"])
header_row = dash.max_row
for c in range(1, 5):
    dash.cell(row=header_row, column=c).font = Font(name="Calibri", bold=True, color="FFFFFF")
    dash.cell(row=header_row, column=c).fill = PatternFill("solid", start_color="2F5597")

# Build sorted category list
cats = sorted({e["category"] for e in DATA})
for cat in cats:
    dash.append([
        cat,
        f'=COUNTIF(\'LeetCode Problems\'!C2:C{last_row_lc}, "{cat}")',
        f'=COUNTIFS(\'LeetCode Problems\'!C2:C{last_row_lc}, "{cat}", \'LeetCode Problems\'!L2:L{last_row_lc}, "<>")',
        f"=IFERROR(C{dash.max_row + 1}/B{dash.max_row + 1}, 0)",
    ])
    dash.cell(row=dash.max_row, column=4).number_format = "0.0%"

# Widths
dash.column_dimensions['A'].width = 28
dash.column_dimensions['B'].width = 12
dash.column_dimensions['C'].width = 14
dash.column_dimensions['D'].width = 10
dash.column_dimensions['E'].width = 10

# ---------------- Coverage Summary ----------------
cov = wb.create_sheet("Coverage Summary")
cov.append(["Category", "Total", "Covered by Sean Prashad", "Covered by Grind 75", "Covered by Both", "Covered by Neither"])
from collections import defaultdict
by_cat = defaultdict(list)
for e in DATA: by_cat[e["category"]].append(e)
for cat in sorted(by_cat.keys()):
    items = by_cat[cat]
    sp = sum(1 for e in items if e["sean_prashad"])
    g = sum(1 for e in items if e["grind75"])
    both = sum(1 for e in items if e["sean_prashad"] and e["grind75"])
    neither = sum(1 for e in items if not e["sean_prashad"] and not e["grind75"])
    cov.append([cat, len(items), sp, g, both, neither])
cov.append([])
cov.append(["TOTAL", len(DATA),
            sum(1 for e in DATA if e["sean_prashad"]),
            sum(1 for e in DATA if e["grind75"]),
            sum(1 for e in DATA if e["sean_prashad"] and e["grind75"]),
            sum(1 for e in DATA if not e["sean_prashad"] and not e["grind75"])])
for col_idx in range(1, 7):
    c = cov.cell(row=1, column=col_idx)
    c.font = header_font; c.fill = header_fill; c.alignment = center
tot = cov.max_row
for col_idx in range(1, 7):
    c = cov.cell(row=tot, column=col_idx)
    c.font = Font(name="Calibri", bold=True)
    c.fill = PatternFill("solid", start_color="D9D9D9")
for i, w in enumerate([26, 10, 26, 22, 22, 22], start=1):
    cov.column_dimensions[get_column_letter(i)].width = w
cov.freeze_panes = "A2"

# ---------------- Read me ----------------
how = wb.create_sheet("Read me")
guide = [
    ("DSA Reference Catalog with LeetCode Mapping", ""),
    ("", ""),
    ("Generated", "2026-05-27"),
    ("Total entries", len(DATA)),
    ("Total LeetCode problems linked", sum(len(e['leetcode']) for e in DATA)),
    ("", ""),
    ("Catalog sheet", "Every data structure / algorithm / technique with definition, complexity, key techniques, Wikipedia link, and a 'LeetCode problems' column listing the 3-7 canonical problems for that technique. Hyperlink on the cell jumps to the first problem."),
    ("LeetCode Problems sheet", "One row per problem (459 total). Filter by Difficulty, Premium, or Entry. Fill in the 'Done?' column (any character) to track progress."),
    ("Progress Dashboard sheet", "Live totals by difficulty and category. Updates automatically as you mark problems done."),
    ("Coverage Summary sheet", "Per-category count of items covered by Sean Prashad's pattern list and Grind 75."),
    ("", ""),
    ("Source: Sean Prashad", "https://seanprashad.com/leetcode-patterns/"),
    ("Source: Grind 75", "https://www.techinterviewhandbook.org/grind75/"),
    ("Source: TIH Cheatsheet", "https://www.techinterviewhandbook.org/algorithms/study-cheatsheet/"),
    ("LeetCode data fetched", "2026-05-27 via LeetCode GraphQL"),
]
for row in guide: how.append(row)
how["A1"].font = Font(name="Calibri", bold=True, size=16, color="2F5597")
how.column_dimensions["A"].width = 34
how.column_dimensions["B"].width = 100
for r in range(1, how.max_row + 1):
    how.cell(row=r, column=1).alignment = Alignment(vertical="top")
    how.cell(row=r, column=2).alignment = Alignment(vertical="top", wrap_text=True)

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  Catalog: {len(DATA)} entries")
print(f"  LeetCode Problems: {sum(len(e['leetcode']) for e in DATA)} rows")
